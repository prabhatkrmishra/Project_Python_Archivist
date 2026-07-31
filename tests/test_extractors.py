from pathlib import Path

import pytest


from archivist.ingestion.extractors import (
    SUPPORTED_EXTENSIONS,
    UnsupportedFileType,
    normalize_text,
    sha256_file,
    should_chunk,
    iter_files,
    extract_csv,
    extract_excel,
    extract_jsonl,
    extract_text,
    chunk_text,
    chunk_csv_by_rows,
    chunk_jsonl_by_lines,
    chunk_code_by_lines,
    cumulative_line_offsets,
)


# Minimal 3-page PDF with no text; enough for page-count based logic.
_MINIMAL_PDF = (
    b"%PDF-1.4\n"
    b"1 0 obj\n<< /Type /Catalog /Pages 2 0 R >>\nendobj\n"
    b"2 0 obj\n<< /Type /Pages /Kids [3 0 R 4 0 R 5 0 R] /Count 3 >>\nendobj\n"
    b"3 0 obj\n<< /Type /Page /Parent 2 0 R /MediaBox [0 0 200 200] >>\nendobj\n"
    b"4 0 obj\n<< /Type /Page /Parent 2 0 R /MediaBox [0 0 200 200] >>\nendobj\n"
    b"5 0 obj\n<< /Type /Page /Parent 2 0 R /MediaBox [0 0 200 200] >>\nendobj\n"
    b"xref\n0 6\n0000000000 65535 f \n0000000009 00000 n \n0000000058 00000 n \n"
    b"0000000115 00000 n \n0000000216 00000 n \n0000000317 00000 n \n"
    b"trailer\n<< /Size 6 /Root 1 0 R >>\nstartxref\n400\n%%EOF\n"
)


def test_normalize_text_lowercases_and_collapses_whitespace():
    assert normalize_text("Hello   WORLD\n\n\tfoo") == "hello world foo"


def test_sha256_file_deterministic(tmp_path: Path):
    f = tmp_path / "a.txt"
    f.write_text("hello")
    assert sha256_file(f) == sha256_file(f)


def test_iter_files_single(tmp_path: Path):
    (tmp_path / "doc.txt").write_text("hi")
    files = list(iter_files(tmp_path / "doc.txt", recursive=False))
    assert len(files) == 1


def test_iter_files_dir(tmp_path: Path):
    (tmp_path / "a.txt").write_text("hi")
    (tmp_path / "b.pdf").write_bytes(b"%PDF-1.4")
    sub = tmp_path / "sub"
    sub.mkdir()
    (sub / "c.docx").write_bytes(b"PK\x03\x04")
    files = sorted(iter_files(tmp_path))
    assert len(files) == 3
    assert files[0].name == "a.txt"
    assert files[1].name == "b.pdf"
    assert files[2].name == "c.docx"


def test_iter_files_ignores_other_extensions(tmp_path: Path):
    (tmp_path / "image.jpg").write_bytes(b"\xff\xd8")
    (tmp_path / "doc.txt").write_text("hi")
    files = list(iter_files(tmp_path))
    assert all(f.suffix == ".txt" for f in files)


def test_should_chunk_small_file(tmp_path: Path):
    (tmp_path / "small.txt").write_text("small content")
    assert not should_chunk(tmp_path / "small.txt", "small content")


# --- CSV Extraction Tests ---


def test_extract_csv_basic(tmp_path: Path):
    f = tmp_path / "data.csv"
    f.write_text("name,age,city\nAlice,30,New York\nBob,25,San Francisco\n")
    result = extract_csv(f)
    assert "name | age | city" in result
    assert "name: Alice | age: 30 | city: New York" in result
    assert "name: Bob | age: 25 | city: San Francisco" in result


def test_extract_csv_tab_delimited(tmp_path: Path):
    f = tmp_path / "data.tsv"
    f.write_text("name\tage\tcity\nAlice\t30\tNew York\nBob\t25\tSan Francisco\n")
    result = extract_csv(f)
    assert "name | age | city" in result
    assert "name: Alice | age: 30 | city: New York" in result


def test_extract_csv_semicolon(tmp_path: Path):
    f = tmp_path / "data.csv"
    f.write_text("name;age;city\nAlice;30;New York\n")
    result = extract_csv(f)
    assert "name | age | city" in result
    assert "name: Alice | age: 30 | city: New York" in result


def test_extract_csv_unicode(tmp_path: Path):
    f = tmp_path / "data.csv"
    f.write_text("name,city\nAlice,café\nBob,résumé\n", encoding="utf-8")
    result = extract_csv(f)
    assert "café" in result
    assert "résumé" in result


def test_extract_csv_empty_rows(tmp_path: Path):
    f = tmp_path / "data.csv"
    f.write_text("name,age\nAlice,30\n\n\nBob,25\n")
    result = extract_csv(f)
    assert "name: Alice | age: 30" in result
    assert "name: Bob | age: 25" in result


def test_extract_csv_single_column(tmp_path: Path):
    f = tmp_path / "data.csv"
    f.write_text("fruit\napple\nbanana\ncherry\n")
    result = extract_csv(f)
    assert "fruit" in result
    assert "apple" in result
    assert "banana" in result


def test_extract_csv_latin1_encoding(tmp_path: Path):
    f = tmp_path / "data.csv"
    f.write_bytes("name,city\nAlice,Ñoño\n".encode("latin-1"))
    result = extract_csv(f)
    assert "Ñoño" in result


# --- Excel Extraction Tests ---


def test_extract_excel_xlsx(tmp_path: Path):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["name", "age", "city"])
    ws.append(["Alice", 30, "New York"])
    ws.append(["Bob", 25, "San Francisco"])
    f = tmp_path / "data.xlsx"
    wb.save(f)
    wb.close()

    result = extract_excel(f)
    assert "[Sheet: Sheet1]" in result
    assert "name | age | city" in result
    assert "name: Alice | age: 30 | city: New York" in result
    assert "name: Bob | age: 25 | city: San Francisco" in result


def test_extract_excel_multiple_sheets(tmp_path: Path):
    from openpyxl import Workbook

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "People"
    ws1.append(["name", "age"])
    ws1.append(["Alice", 30])
    ws2 = wb.create_sheet("Cities")
    ws2.append(["city", "population"])
    ws2.append(["NYC", 8000000])
    f = tmp_path / "multi.xlsx"
    wb.save(f)
    wb.close()

    result = extract_excel(f)
    assert "[Sheet: People]" in result
    assert "[Sheet: Cities]" in result
    assert "name: Alice | age: 30" in result
    assert "city: NYC | population: 8000000" in result


def test_extract_excel_xls(tmp_path: Path):
    import xlwt

    wb = xlwt.Workbook()
    ws = wb.add_sheet("Data")
    ws.write(0, 0, "name")
    ws.write(0, 1, "score")
    ws.write(1, 0, "Alice")
    ws.write(1, 1, 95)
    f = tmp_path / "data.xls"
    wb.save(f)

    result = extract_excel(f)
    assert "[Sheet: Data]" in result
    assert "name | score" in result
    assert "name: Alice | score: 95.0" in result


def test_extract_excel_xls_empty_sheet(tmp_path: Path):
    """An .xls sheet with zero rows is skipped, not emitted as an empty header."""
    import xlwt

    wb = xlwt.Workbook()
    wb.add_sheet("Empty")
    f = tmp_path / "empty.xls"
    wb.save(f)

    assert extract_excel(f) == ""


def test_extract_excel_empty_sheet(tmp_path: Path):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Empty"
    f = tmp_path / "empty.xlsx"
    wb.save(f)
    wb.close()

    result = extract_excel(f)
    assert result == ""


# --- JSONL Extraction Tests ---


def test_extract_jsonl_basic(tmp_path: Path):
    f = tmp_path / "data.jsonl"
    f.write_text('{"name": "Alice", "age": 30}\n{"name": "Bob", "age": 25}\n')
    result = extract_jsonl(f)
    assert "L1:" in result
    assert "name: Alice" in result
    assert "age: 30" in result
    assert "L2:" in result
    assert "name: Bob" in result


def test_extract_jsonl_nested(tmp_path: Path):
    f = tmp_path / "data.jsonl"
    f.write_text('{"name": "Alice", "address": {"city": "NYC", "zip": "10001"}}\n')
    result = extract_jsonl(f)
    assert "address.city: NYC" in result
    assert "address.zip: 10001" in result


def test_extract_jsonl_arrays(tmp_path: Path):
    f = tmp_path / "data.jsonl"
    f.write_text('{"name": "Alice", "tags": ["admin", "user"]}\n')
    result = extract_jsonl(f)
    assert "tags:" in result
    assert "admin" in result


def test_extract_jsonl_empty_lines(tmp_path: Path):
    f = tmp_path / "data.jsonl"
    f.write_text('{"name": "Alice"}\n\n\n{"name": "Bob"}\n')
    result = extract_jsonl(f)
    assert "L1:" in result
    assert "L4:" in result


def test_extract_jsonl_invalid_json(tmp_path: Path):
    f = tmp_path / "data.jsonl"
    f.write_text('{"name": "Alice"}\nnot valid json\n{"name": "Bob"}\n')
    result = extract_jsonl(f)
    assert "L1:" in result
    assert "L2: not valid json" in result
    assert "L3:" in result


def test_extract_jsonl_mixed_types(tmp_path: Path):
    f = tmp_path / "data.jsonl"
    f.write_text('"just a string"\n42\n{"key": "value"}\n')
    result = extract_jsonl(f)
    assert "L1: just a string" in result
    assert "L2: 42" in result
    assert "L3: key: value" in result


def test_extract_jsonl_list_line(tmp_path: Path):
    """A JSON array line is dumped as-is rather than flattened."""
    f = tmp_path / "data.jsonl"
    f.write_text('[1, 2, "three"]\n')
    result = extract_jsonl(f)
    assert 'L1: [1, 2, "three"]' in result


def test_read_file_with_fallback_all_encodings_fail():
    """Bytes rejected by every encoding fall through to errors='replace'."""
    import archivist.ingestion.extractors as ex

    class WeirdPath:
        def read_text(self, encoding=None, errors=None):
            if errors == "replace":
                return "fallback content"
            raise UnicodeDecodeError(encoding, b"\xff", 0, 1, "invalid byte")

    assert ex._read_file_with_fallback(WeirdPath()) == "fallback content"


# --- Chunking Tests ---


def test_chunk_csv_by_rows():
    lines = ["name,age,city"] + [f"person{i},{20+i},city{i}" for i in range(100)]
    text = "\n".join(lines)
    chunks = chunk_csv_by_rows(text, rows_per_chunk=30)
    assert len(chunks) == 4  # 30+30+30+10
    assert chunks[0].startswith("name,age,city\n")
    assert "person0" in chunks[0]
    assert "person29" in chunks[0]
    assert "person30" in chunks[1]


def test_chunk_jsonl_by_lines():
    lines = [f'{{"id": {i}, "val": "x"}}' for i in range(1200)]
    text = "\n".join(lines)
    chunks = chunk_jsonl_by_lines(text, lines_per_chunk=500)
    assert len(chunks) == 3  # 500+500+200


def test_chunk_csv_single_chunk():
    text = "name,age\nAlice,30\nBob,25\n"
    chunks = chunk_csv_by_rows(text, rows_per_chunk=100)
    assert len(chunks) == 1
    assert chunks[0] == text


# --- File Discovery Tests ---


def test_iter_files_includes_xlsx_jsonl(tmp_path: Path):
    (tmp_path / "a.xlsx").write_bytes(b"")
    (tmp_path / "b.jsonl").write_text("")
    (tmp_path / "c.txt").write_text("")
    files = sorted(iter_files(tmp_path))
    names = [f.name for f in files]
    assert "a.xlsx" in names
    assert "b.jsonl" in names
    assert "c.txt" in names


def test_supported_extensions_includes_tabular():
    assert ".csv" in SUPPORTED_EXTENSIONS
    assert ".tsv" in SUPPORTED_EXTENSIONS
    assert ".xls" in SUPPORTED_EXTENSIONS
    assert ".xlsx" in SUPPORTED_EXTENSIONS
    assert ".jsonl" in SUPPORTED_EXTENSIONS


# --- Code Chunking Tests ---


def test_chunk_code_by_lines_small_file(tmp_path: Path):
    """Small files stay as single chunk."""
    text = "def hello():\n    return 'world'\n"
    chunks = chunk_code_by_lines(text)
    assert len(chunks) == 1
    assert chunks[0] == text


def test_chunk_code_by_lines_exact_boundary(tmp_path: Path):
    """File at exactly 1500 lines stays as single chunk."""
    lines = [f"L{i:04d}" for i in range(1500)]
    text = "\n".join(lines)
    chunks = chunk_code_by_lines(text)
    assert len(chunks) == 1


def test_chunk_code_by_lines_splits_large_file(tmp_path: Path):
    """Large files are split into chunks of 1500 lines."""
    lines = [f"L{i:04d}" for i in range(4200)]
    text = "\n".join(lines)
    chunks = chunk_code_by_lines(text)
    assert len(chunks) == 3  # 1500+1500+1200
    for c in chunks:
        assert c.strip()  # no empty chunks


def test_chunk_code_by_lines_preserves_content(tmp_path: Path):
    """Chunk content matches original lines."""
    lines = [f"line{i}" for i in range(3200)]
    text = "\n".join(lines)
    chunks = chunk_code_by_lines(text)
    assert len(chunks) == 3  # 1500+1500+200
    # First chunk has first 1500 lines
    chunk1_lines = chunks[0].split("\n")
    assert len(chunk1_lines) == 1500
    assert chunk1_lines[0] == "line0"
    assert chunk1_lines[1499] == "line1499"
    # Second chunk has next 1500 lines
    chunk2_lines = chunks[1].split("\n")
    assert len(chunk2_lines) == 1500
    assert chunk2_lines[0] == "line1500"

# --- Edge cases: delimiter fallback, empty inputs, unsupported types ---


def test_extract_csv_delimiter_fallback(tmp_path):
    """Sniffer fails on freeform text; extraction falls back to comma default."""
    f = tmp_path / "freeform.csv"
    f.write_text("just some freeform text\nwith a second line\n")
    result = extract_csv(f)
    assert "just some freeform text" in result


def test_extract_csv_empty_header_cell(tmp_path):
    f = tmp_path / "data.csv"
    f.write_text(",city\nAlice,NYC\n")
    result = extract_csv(f)
    assert "Alice" in result


def test_extract_csv_empty_file(tmp_path):
    f = tmp_path / "empty.csv"
    f.write_text("")
    assert extract_csv(f) == ""


def test_extract_csv_only_blank_rows(tmp_path):
    f = tmp_path / "blank.csv"
    f.write_text("\n\n\n")
    assert extract_csv(f) == ""


def test_extract_text_unsupported_extension(tmp_path):
    f = tmp_path / "image.jpg"
    f.write_bytes(b"\xff\xd8\xff\xe0")
    with pytest.raises(UnsupportedFileType):
        extract_text(f)


def test_chunk_text_unknown_type_returns_single(tmp_path):
    # .xlsx is supported for extraction but has no dedicated chunker.
    f = tmp_path / "data.xlsx"
    f.write_bytes(b"")
    assert chunk_text(f, "some content") == ["some content"]


def test_chunk_csv_by_rows_empty_text():
    assert chunk_csv_by_rows("") == [""]


def test_chunk_csv_by_rows_header_only():
    assert chunk_csv_by_rows("name,age") == ["name,age"]


def test_chunk_jsonl_by_lines_empty_text():
    assert chunk_jsonl_by_lines("") == [""]


def test_iter_files_finds_extensionless_build_files(tmp_path):
    (tmp_path / "Makefile").write_text("all:\n\techo hi\n")
    (tmp_path / "Dockerfile").write_text("FROM python\n")
    files = list(iter_files(tmp_path))
    assert {f.name for f in files} == {"Makefile", "Dockerfile"}


# --- should_chunk thresholds (monkeypatched to keep fixtures small) ---


def test_should_chunk_code_over_line_threshold(tmp_path, monkeypatch):
    import archivist.ingestion.extractors as ex

    monkeypatch.setattr(ex, "_CHUNK_THRESHOLD_BYTES", 10 * 1024 * 1024)
    f = tmp_path / "big.py"
    f.write_text("\n".join(f"line{i}" for i in range(501)))
    assert should_chunk(f, f.read_text())


def test_should_chunk_csv_over_row_threshold(tmp_path, monkeypatch):
    import archivist.ingestion.extractors as ex

    monkeypatch.setattr(ex, "_CHUNK_THRESHOLD_BYTES", 10 * 1024 * 1024)
    f = tmp_path / "big.csv"
    f.write_text("\n".join([f"row{i}" for i in range(5001)]))
    assert should_chunk(f, f.read_text())


def test_should_chunk_jsonl_over_line_threshold(tmp_path, monkeypatch):
    import archivist.ingestion.extractors as ex

    monkeypatch.setattr(ex, "_CHUNK_THRESHOLD_BYTES", 10 * 1024 * 1024)
    f = tmp_path / "big.jsonl"
    f.write_text("\n".join([f'{{"i": {i}}}' for i in range(10001)]))
    assert should_chunk(f, f.read_text())


def test_should_chunk_pdf_over_page_threshold(tmp_path, monkeypatch):
    import archivist.ingestion.extractors as ex

    monkeypatch.setattr(ex, "_CHUNK_THRESHOLD_PAGES", 2)
    f = tmp_path / "pages.pdf"
    f.write_bytes(_MINIMAL_PDF)  # 3 pages
    assert should_chunk(f, "")


def test_should_chunk_pdf_reader_error_does_not_raise(tmp_path, monkeypatch):
    """A PdfReader failure is swallowed: size-based chunking still applies."""
    import archivist.ingestion.extractors as ex

    def boom(path):
        raise RuntimeError("pdf read failed")

    monkeypatch.setattr(ex.pypdf, "PdfReader", boom)
    f = tmp_path / "pages.pdf"
    f.write_bytes(_MINIMAL_PDF)
    assert should_chunk(f, "") is False


# --- cumulative line offsets ---


def test_cumulative_line_offsets_variable_chunk_sizes():
    chunks = ["a\nb\nc", "d\ne", "f"]
    assert cumulative_line_offsets(chunks) == [0, 3, 5]


def test_cumulative_line_offsets_empty_chunks():
    assert cumulative_line_offsets([]) == []
    assert cumulative_line_offsets(["", "x"]) == [0, 1]
